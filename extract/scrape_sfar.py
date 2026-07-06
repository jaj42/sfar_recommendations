#!/usr/bin/env python3
"""
SFAR Recommendations PDF Scraper

Discovers every SFAR clinical recommendation from the recommendations index,
resolves each entry to its concrete PDF (direct upload or WP Download Manager
endpoint), downloads the PDFs, and writes a manifest XLSX for human curation.

Discovery reads a local snapshot of the index page ("Recommandations - La SFAR.html")
by default: it is complete (includes 2026), reproducible, and avoids Cloudflare 403s.
PDFs themselves are always fetched from the live site.

Usage:
    python scrape_sfar.py                 # discover + resolve + download + verify
    python scrape_sfar.py --dry-run       # discover + resolve + manifest, no download
    python scrape_sfar.py --verify-only   # re-run the completeness check on an existing manifest
    python scrape_sfar.py --help          # all options
"""

import argparse
import hashlib
import json
import re
import sys
import threading
import time
import unicodedata
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urljoin, urlparse, unquote

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #

BASE_URL = "https://sfar.org"
RECOMMENDATIONS_URL = "https://sfar.org/recommandations/"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INDEX_HTML = SCRIPT_DIR / "Recommandations - La SFAR.html"
REFERENCE_MD = SCRIPT_DIR / "Recommandations - La SFAR.md"

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,image/apng,*/*;q=0.8"
    ),
    "Accept-Language": "fr-FR,fr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": RECOMMENDATIONS_URL,
    "DNT": "1",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

# tab id -> discipline label, keyed by the aria-labelledby value on div.tab-pane
TAB_DISCIPLINE = {
    "chronologie": "Chronologie",
    "anesth": "Anesthesie",
    "réanimation": "Reanimation",
    "reanimation": "Reanimation",
    "urgence": "Urgences",
}

# tr class -> best-effort recommendation type
ROW_CLASS_TYPE = {
    "table-rfe-border": "RFE",
    "table-rpp-border": "RPP",
    "table-preco-border": "Preconisation",
}

# td cell class -> discipline refinement (chronologie rows also carry -anesth etc.)
CELL_CLASS_DISCIPLINE = {
    "td-border-rfe-anesth": "Anesthesie",
    "td-border-rfe-rea": "Reanimation",
    "td-border-rfe-urgence": "Urgences",
}

# Footer / boilerplate files that appear on every landing page and must be dropped.
NOISE_PATTERNS = [
    r"\bcgv\b",
    r"certificat",
    r"reglement",
    r"r[eè]glement",
    r"\bcharte\b",
    r"\bagenda\b",
    r"\bnewsletter\b",
    r"info-sfar-",
    r"mentions-legales",
]

WPDMDL_RE = re.compile(r"wpdmdl=(\d+)")

# Document file extensions we treat as real recommendation attachments (PDFs
# plus Office documents — some recos ship as .docx/.xlsx). Images are excluded
# on purpose: the only image-hosted "recos" on the site are low-value scans,
# not documents, so they are reported as misses rather than downloaded.
DOC_EXTS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls", ".odt", ".rtf"}

# Extensions accepted as a direct target straight from the index.
INDEX_FILE_EXTS = DOC_EXTS

# Content-Type -> extension, used to name files that lack one (e.g. the wpdmdl
# and bare /download/ endpoints stream a file with no extension in the URL).
CONTENT_TYPE_EXT = {
    "application/pdf": ".pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/msword": ".doc",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": ".pptx",
    "application/vnd.ms-powerpoint": ".ppt",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/vnd.ms-excel": ".xls",
    "application/vnd.oasis.opendocument.text": ".odt",
    "application/rtf": ".rtf",
    "text/rtf": ".rtf",
}
# Content-Types that indicate a real document (used to accept a landing page
# that redirects straight to the file rather than to an HTML page).
DOC_CONTENT_TYPES = set(CONTENT_TYPE_EXT)
KNOWN_EXTS = set(CONTENT_TYPE_EXT.values()) | INDEX_FILE_EXTS


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #

def strip_accents(text):
    text = unicodedata.normalize("NFKD", text)
    return "".join(c for c in text if not unicodedata.combining(c))


def is_noise_file(url):
    """True if the URL is boilerplate (CGV, reglement, certificat, ...)."""
    norm = strip_accents(unquote(url).lower())
    return any(re.search(p, norm) for p in NOISE_PATTERNS)


def url_ext(url):
    """Lower-case file extension of a URL's path (query/fragment stripped)."""
    return Path(urlparse(url).path).suffix.lower()


def ensure_extension(filename, content_type):
    """Ensure ``filename`` carries a sensible extension.

    Keep any recognised extension already present; otherwise append one derived
    from the Content-Type. Falls back to ``.pdf`` only when nothing is known
    (the overwhelming majority of SFAR documents are PDFs).
    """
    filename = (filename or "").strip() or "document"
    if Path(filename).suffix.lower() in KNOWN_EXTS:
        return filename
    ext = CONTENT_TYPE_EXT.get(content_type, "")
    if ext:
        return filename + ext
    # Unknown content-type and no usable extension: default to .pdf.
    return filename if Path(filename).suffix else filename + ".pdf"


def unwrap_urldefense(url):
    """Unwrap a Proofpoint urldefense.com/v3/__<REAL>__;... link to <REAL>.

    SFAR sometimes links partner-society PDFs (e.g. urofrance.org) through this
    wrapper, which hides the real URL and its .pdf extension.
    """
    m = re.match(r"https?://urldefense\.com/v3/__(.+?)__;", url)
    return unquote(m.group(1)) if m else url


def apply_has_type(base_type, title, download_url):
    """Upgrade the type to 'HAS' for Haute Autorité de Santé recommendations.

    Signalled either by the document being hosted on has-sante.fr or by an
    explicit 'HAS' token in the title (e.g. '... - HAS 2025').
    """
    if "has-sante.fr" in urlparse(download_url).netloc:
        return "HAS"
    if re.search(r"\bHAS\b", title):
        return "HAS"
    return base_type


def clean_wpdmdl_url(url):
    """Normalise a WP Download Manager URL to /download/<slug>/?wpdmdl=<id>.

    Drops the volatile &refresh=... nonce so the URL is stable for dedup while
    still serving the PDF.
    """
    m = WPDMDL_RE.search(url)
    if not m:
        return url
    parsed = urlparse(url)
    base = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
    return f"{base}?wpdmdl={m.group(1)}"


def absolutize(href, base=BASE_URL):
    href = href.strip()
    return urljoin(base, href)


def filename_from_disposition(header):
    """Extract filename from a Content-Disposition header, or None."""
    if not header:
        return None
    m = re.search(r"filename\*?=(?:UTF-8''|\")?([^\";]+)", header, re.IGNORECASE)
    if not m:
        return None
    name = m.group(1).strip().strip('"')
    return Path(name).name or None


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# --------------------------------------------------------------------------- #
# Scraper
# --------------------------------------------------------------------------- #

class SFARScraper:
    def __init__(self, output_dir, index_html=None, live_index=False,
                 workers=6, delay=0.5):
        self.output_dir = Path(output_dir)
        self.pdfs_dir = self.output_dir / "pdfs"
        self.pdfs_dir.mkdir(parents=True, exist_ok=True)

        self.index_html = Path(index_html) if index_html else DEFAULT_INDEX_HTML
        self.live_index = live_index
        self.workers = max(1, workers)
        self.delay = delay

        self.session = requests.Session()
        self.session.headers.update(BROWSER_HEADERS)

        self._throttle_lock = threading.Lock()
        self._last_request = 0.0

        self.existing_pdfs = self._load_existing_pdfs()
        # Records from a previous run (key -> doc), used to skip already-present
        # files on re-runs without issuing any network request.
        self.prev_records = self._load_prev_records()

    # -- fetching ---------------------------------------------------------- #

    def _throttle(self):
        with self._throttle_lock:
            wait = self.delay - (time.time() - self._last_request)
            if wait > 0:
                time.sleep(wait)
            self._last_request = time.time()

    def fetch_response(self, url, max_retries=3):
        """GET a URL returning the Response (or None). Retries with backoff."""
        for attempt in range(max_retries):
            self._throttle()
            try:
                r = self.session.get(url, timeout=30)
                if r.status_code == 200:
                    return r
                if r.status_code == 403:
                    print(f"  ⛔ 403 Forbidden: {url}")
                    return r
                if attempt == max_retries - 1:
                    return r
            except requests.RequestException as e:
                if attempt == max_retries - 1:
                    print(f"  ❌ fetch failed {url}: {e}")
                    return None
            time.sleep(2 ** attempt)
        return None

    def fetch_text(self, url, max_retries=3):
        """GET a URL returning (text, status). Retries with backoff."""
        r = self.fetch_response(url, max_retries)
        if r is None:
            return None, 0
        if r.status_code == 200:
            return r.text, 200
        return None, r.status_code

    def _load_existing_pdfs(self):
        """Basenames of the 87 curated PDFs in ../pdfs (for coverage diff)."""
        names = set()
        parent = SCRIPT_DIR.parent / "pdfs"
        if parent.exists():
            for p in parent.rglob("*.pdf"):
                names.add(p.name.lower())
        return names

    def _load_prev_records(self):
        """Load {key: doc} from a previous run's discovery.json, if present.

        Only successful downloads (those with a recorded filename) are kept, so a
        re-run can skip a file that is already on disk without re-fetching it.
        """
        path = self.output_dir / "discovery.json"
        records = {}
        if path.exists():
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
                for d in data.get("documents", []):
                    if d.get("key") and d.get("filename"):
                        records[d["key"]] = d
            except (json.JSONDecodeError, OSError):
                pass
        return records

    # -- discovery --------------------------------------------------------- #

    def _load_index_html(self):
        if self.live_index:
            print(f"🌐 Fetching live index: {RECOMMENDATIONS_URL}")
            html, status = self.fetch_text(RECOMMENDATIONS_URL)
            if status == 200 and html:
                return html
            print(f"⚠️  Live index fetch failed (status {status}); "
                  f"falling back to local snapshot.")
        if not self.index_html.exists():
            print(f"ERROR: index HTML not found: {self.index_html}")
            return None
        print(f"📄 Reading index snapshot: {self.index_html.name}")
        return self.index_html.read_text(encoding="utf-8")

    @staticmethod
    def _discipline_for_tab(tab_pane):
        label = (tab_pane.get("aria-labelledby", "") or tab_pane.get("id", "")).lower()
        for key, disc in TAB_DISCIPLINE.items():
            if key in label:
                return disc
        return "Unknown"

    @staticmethod
    def _type_from_row(tr):
        for cls in tr.get("class", []):
            if cls in ROW_CLASS_TYPE:
                return ROW_CLASS_TYPE[cls]
        return "Unknown"

    @staticmethod
    def _discipline_from_cell(td):
        for cls in td.get("class", []):
            if cls in CELL_CLASS_DISCIPLINE:
                return CELL_CLASS_DISCIPLINE[cls]
        return None

    def discover(self):
        """Parse the index HTML into a list of row entries."""
        html = self._load_index_html()
        if not html:
            return []
        soup = BeautifulSoup(html, "lxml")

        entries = []
        for tab_pane in soup.find_all("div", class_="tab-pane"):
            tab_discipline = self._discipline_for_tab(tab_pane)
            is_chrono = tab_discipline == "Chronologie"

            for panel in tab_pane.find_all("div", class_="panel-default"):
                heading_tag = panel.find("h4")
                heading = heading_tag.get_text(strip=True) if heading_tag else ""

                year = None
                subdomain = None
                if is_chrono:
                    m = re.search(r"\b(19|20)\d{2}\b", heading)
                    year = int(m.group()) if m else None
                else:
                    subdomain = heading or None

                for table in panel.find_all("table"):
                    for tr in table.find_all("tr"):
                        link = tr.find("a", href=True)
                        if not link:
                            continue
                        href = link["href"].strip()
                        if not href or href.startswith("#"):
                            continue
                        title = link.get_text(strip=True)
                        if not title:
                            continue

                        first_td = tr.find("td")
                        cell_discipline = (
                            self._discipline_from_cell(first_td) if first_td else None
                        )
                        discipline = cell_discipline or (
                            None if is_chrono else tab_discipline
                        )

                        entries.append({
                            "title": title,
                            "href": absolutize(href),
                            "type": self._type_from_row(tr),
                            "discipline": discipline,
                            "subdomain": subdomain,
                            "year": year,
                            "source_tab": tab_discipline,
                        })

        print(f"📊 Discovered {len(entries)} index rows "
              f"({len({e['href'] for e in entries})} unique hrefs)")
        return entries

    # -- resolution -------------------------------------------------------- #

    def _targets_from_landing(self, url):
        """Fetch a landing page and return download-target URLs found in it.

        WP Download Manager renders a button
        `<a class="wpdm-download-link" href="#"
            data-downloadurl=".../download/<slug>/?wpdmdl=<id>&refresh=...">`;
        the real PDF URL lives in ``data-downloadurl`` (the visible ``href`` is
        just ``#`` or the package landing). Older recos instead link a direct
        ``wp-content/uploads/.../*.pdf``.
        """
        r = self.fetch_response(url)
        if r is None:
            return [], 0
        if r.status_code != 200:
            return [], r.status_code

        ctype = r.headers.get("content-type", "").split(";")[0].strip().lower()

        # The "landing page" sometimes redirects straight to the document itself
        # (a PDF- or Office-hosted reco). If the response *is* a document, its
        # final URL is the download target. Non-document files (e.g. an image
        # scan) are intentionally ignored — they are not usable recommendations.
        if ctype in DOC_CONTENT_TYPES or url_ext(r.url) in DOC_EXTS:
            return [r.url], 200

        soup = BeautifulSoup(r.text, "lxml")

        targets = []
        seen = set()

        def add(full):
            if full not in seen:
                seen.add(full)
                targets.append(full)

        # WP Download Manager buttons carry the real URL in data-downloadurl.
        for el in soup.find_all(attrs={"data-downloadurl": True}):
            raw = el["data-downloadurl"].strip()
            if "wpdmdl=" in raw:
                add(clean_wpdmdl_url(absolutize(raw, url)))

        # Any anchor already carrying an explicit ?wpdmdl= (belt and braces),
        # plus direct document files (unwrapping Proofpoint urldefense links
        # first). Images are excluded here — they are usually page decoration.
        for a in soup.find_all("a", href=True):
            href = unwrap_urldefense(a["href"].strip())
            if "wpdmdl=" in href:
                add(clean_wpdmdl_url(absolutize(href, url)))
            elif url_ext(href) in DOC_EXTS:
                # Any document the landing page points to, on any domain (SFAR
                # often links HAS-hosted PDFs directly); boilerplate is dropped.
                if not is_noise_file(href):
                    add(absolutize(href, url))

        return targets, 200

    def resolve(self, entry):
        """
        Resolve an index entry into a list of download targets.

        Each target: {download_url, wpdmdl, landing_url}.
        """
        href = unwrap_urldefense(entry["href"])
        parsed = urlparse(href)

        def make(url, landing=None):
            url = clean_wpdmdl_url(url)
            m = WPDMDL_RE.search(url)
            return {
                "download_url": url,
                "wpdmdl": m.group(1) if m else None,
                "landing_url": landing,
            }

        # Case 1: direct file link in the index (PDF, Office doc, image, ...)
        if url_ext(href) in INDEX_FILE_EXTS:
            if is_noise_file(href):
                return []
            return [make(href)]

        # Case 2: WP Download Manager endpoint carrying an explicit wpdmdl id.
        if "wpdmdl=" in (parsed.query or ""):
            return [make(href)]

        # A bare /download/<slug>/ (no id) serves HTML, not the PDF, so treat it
        # like a landing page and pull the real target(s) from data-downloadurl.

        # Case 3: landing page -> fetch and extract targets
        targets, status = self._targets_from_landing(href)
        entry["_landing_status"] = status
        return [make(t, landing=href) for t in targets]

    def resolve_all(self, entries):
        """Resolve every entry concurrently and dedup by target identity."""
        print(f"🔗 Resolving {len(entries)} entries "
              f"({self.workers} workers)...")

        def work(idx_entry):
            idx, entry = idx_entry
            try:
                targets = self.resolve(entry)
            except Exception as e:  # noqa: BLE001
                print(f"  ❌ resolve error [{entry['title'][:40]}]: {e}")
                targets = []
            return entry, targets

        resolved = []  # (entry, target) pairs
        unresolved = []  # entries that produced no target
        with ThreadPoolExecutor(max_workers=self.workers) as pool:
            futures = [pool.submit(work, ie) for ie in enumerate(entries)]
            done = 0
            for fut in as_completed(futures):
                entry, targets = fut.result()
                done += 1
                if not targets:
                    unresolved.append(entry)
                for t in targets:
                    resolved.append((entry, t))
                if done % 25 == 0 or done == len(entries):
                    print(f"  ... resolved {done}/{len(entries)}")

        print(f"✅ Resolved {len(resolved)} targets from "
              f"{len(entries) - len(unresolved)} entries; "
              f"{len(unresolved)} entries produced no target.")

        # Dedup by identity key: wpdmdl id, else absolute download URL.
        docs = {}
        for entry, target in resolved:
            key = f"wpdmdl:{target['wpdmdl']}" if target["wpdmdl"] else target["download_url"]
            if key not in docs:
                docs[key] = {
                    "key": key,
                    "download_url": target["download_url"],
                    "wpdmdl": target["wpdmdl"],
                    "landing_url": target["landing_url"],
                    "title": entry["title"],
                    "type": apply_has_type(entry["type"], entry["title"],
                                           target["download_url"]),
                    "discipline": entry["discipline"],
                    "subdomain": entry["subdomain"],
                    "year": entry["year"],
                    "source_tabs": {entry["source_tab"]},
                }
            else:
                self._merge_context(docs[key], entry)

        print(f"🎯 Deduplicated to {len(docs)} unique documents.")
        return list(docs.values()), unresolved

    @staticmethod
    def _merge_context(doc, entry):
        doc["source_tabs"].add(entry["source_tab"])
        if not doc.get("year") and entry.get("year"):
            doc["year"] = entry["year"]
        for field in ("discipline", "subdomain"):
            new = entry.get(field)
            cur = doc.get(field)
            if new and new not in (cur or ""):
                doc[field] = new if not cur else f"{cur} | {new}"
        if doc["type"] in (None, "Unknown") and entry["type"] not in (None, "Unknown"):
            doc["type"] = entry["type"]

    # -- download ---------------------------------------------------------- #

    def download(self, doc, html_retries=3):
        """Download one document; annotate doc in place with results.

        WP Download Manager occasionally serves its "Fichier non trouvé" HTML
        interstitial instead of the file under concurrent load; that same URL
        then streams the file fine on a sequential retry. So an HTML response is
        retried a few times with backoff before being treated as a genuine
        upstream-missing failure.
        """
        url = doc["download_url"]
        year = doc.get("year") or "unknown"
        year_dir = self.pdfs_dir / str(year)

        # Skip entirely (no network) if a previous run already saved this file.
        prev = self.prev_records.get(doc.get("key"))
        if prev and prev.get("filename"):
            prev_dest = (self.pdfs_dir / str(prev.get("year") or "unknown")
                         / prev["filename"])
            if prev_dest.exists() and prev_dest.stat().st_size > 0:
                doc.update(error="", downloaded=False,
                           bytes=prev_dest.stat().st_size,
                           sha256=prev.get("sha256") or sha256_file(prev_dest),
                           filename=prev["filename"],
                           content_type=prev.get("content_type", ""),
                           skipped=True)
                return doc

        try:
            for attempt in range(html_retries + 1):
                self._throttle()
                with self.session.get(url, stream=True, timeout=60) as r:
                    if r.status_code != 200:
                        doc.update(error=f"HTTP {r.status_code}", downloaded=False,
                                   bytes=0, sha256="", filename="",
                                   content_type=r.headers.get("content-type", ""))
                        return doc

                    ctype = r.headers.get("content-type", "").split(";")[0].strip().lower()

                    # HTML from a file endpoint is not the document. Retry (it is
                    # often a transient WPDM interstitial); if it persists, the
                    # file is genuinely missing/removed upstream.
                    if ctype == "text/html" or "xml" in ctype:
                        if attempt < html_retries:
                            time.sleep(1.0 + attempt)
                            continue
                        doc.update(error="source returned HTML (file missing upstream?)",
                                   downloaded=False, bytes=0, sha256="", filename="",
                                   content_type=ctype)
                        return doc

                    # Not a document (e.g. an image scan): don't save it.
                    if ctype.startswith("image/") or (
                        ctype and ctype not in DOC_CONTENT_TYPES
                        and ctype != "application/octet-stream"
                        and url_ext(r.url) not in DOC_EXTS
                    ):
                        doc.update(error=f"non-document content-type: {ctype}",
                                   downloaded=False, bytes=0, sha256="", filename="",
                                   content_type=ctype)
                        return doc

                    return self._save_stream(r, doc, year_dir, ctype)
        except requests.RequestException as e:
            doc.update(error=str(e), downloaded=False, bytes=0, sha256="",
                       filename="", content_type="")
            return doc

    def _save_stream(self, r, doc, year_dir, ctype):
        """Persist a streaming Response body; annotate and return doc."""
        filename = ensure_extension(
            filename_from_disposition(r.headers.get("content-disposition"))
            or Path(urlparse(r.url).path).name
            or str(doc.get("wpdmdl") or "document"),
            ctype,
        )

        year_dir.mkdir(parents=True, exist_ok=True)
        dest = year_dir / filename

        # Idempotency: skip if a same-named file already exists non-empty.
        if dest.exists() and dest.stat().st_size > 0:
            doc.update(error="", downloaded=False,
                       bytes=dest.stat().st_size,
                       sha256=sha256_file(dest), filename=filename,
                       content_type=ctype)
            return doc

        size = 0
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=65536):
                if chunk:
                    f.write(chunk)
                    size += len(chunk)

        error = "empty download" if size == 0 else ""
        doc.update(error=error, downloaded=True, bytes=size,
                   sha256=sha256_file(dest), filename=filename,
                   content_type=ctype)
        return doc

    def download_all(self, docs):
        print(f"📥 Downloading {len(docs)} documents ({self.workers} workers)...")
        with ThreadPoolExecutor(max_workers=self.workers) as pool:
            futures = [pool.submit(self.download, d) for d in docs]
            done = 0
            for fut in as_completed(futures):
                doc = fut.result()
                done += 1
                status = "✔" if not doc.get("error") else f"✗ {doc['error']}"
                tag = "new" if doc.get("downloaded") else "skip"
                print(f"  [{done}/{len(docs)}] {status} ({tag}) "
                      f"{doc.get('filename', '')[:60]}")
        return docs

    # -- outputs ----------------------------------------------------------- #

    def write_discovery_json(self, docs, unresolved):
        payload = {
            "documents": [
                {**d, "source_tabs": sorted(d.get("source_tabs", []))} for d in docs
            ],
            "unresolved": unresolved,
        }
        path = self.output_dir / "discovery.json"
        path.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                        encoding="utf-8")
        print(f"📋 Wrote {path}")

    # Manifest columns: (header, key-or-callable, width, numeric?)
    MANIFEST_COLUMNS = [
        ("type", lambda d: d.get("type", ""), 14, False),
        ("discipline", lambda d: d.get("discipline") or "", 16, False),
        ("subdomain", lambda d: d.get("subdomain") or "", 18, False),
        ("year", lambda d: d.get("year"), 8, True),
        ("title", lambda d: d.get("title", ""), 70, False),
        ("landing_url", lambda d: d.get("landing_url") or "", 45, False),
        ("download_url", lambda d: d.get("download_url", ""), 45, False),
        ("wpdmdl", lambda d: d.get("wpdmdl") or "", 12, False),
        ("filename", lambda d: d.get("filename") or "", 40, False),
        ("bytes", lambda d: d.get("bytes"), 12, True),
        ("sha256", lambda d: d.get("sha256") or "", 24, False),
        ("is_new", lambda d: 1 if d.get("downloaded") else 0, 8, True),
        ("include", lambda d: "", 10, False),
    ]

    def write_manifest(self, docs):
        path = self.output_dir / "manifest.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "manifest"

        base_font = Font(name="Arial", size=10)
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", start_color="305496")
        wrap = Alignment(vertical="top", wrap_text=True)

        headers = [c[0] for c in self.MANIFEST_COLUMNS]
        ws.append(headers)
        for col, (_, _, width, _) in enumerate(self.MANIFEST_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width

        # Sort for a readable manifest: newest year first, then type, then title.
        ordered = sorted(
            docs,
            key=lambda d: (-(d.get("year") or 0), str(d.get("type")), d.get("title", "")),
        )
        for d in ordered:
            row = []
            for _, getter, _, numeric in self.MANIFEST_COLUMNS:
                val = getter(d)
                if numeric and (val is None or val == ""):
                    val = None
                row.append(val)
            ws.append(row)

        for r in range(2, ws.max_row + 1):
            for c in range(1, len(self.MANIFEST_COLUMNS) + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = base_font
                if self.MANIFEST_COLUMNS[c - 1][0] == "title":
                    cell.alignment = wrap
                elif self.MANIFEST_COLUMNS[c - 1][0] == "year":
                    cell.number_format = "0"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.MANIFEST_COLUMNS))}{ws.max_row}"
        wb.save(path)
        print(f"📝 Wrote {path} ({len(docs)} rows)")

    # -- verification ------------------------------------------------------ #

    def verify_complete(self, docs, unresolved, downloaded=True):
        """
        Cross-check completeness. Returns True if comprehensive.
        Writes completeness_report.txt and prints a summary.
        """
        lines = []

        def out(s=""):
            print(s)
            lines.append(s)

        out("=" * 60)
        out("COMPLETENESS REPORT")
        out("=" * 60)
        out(f"Unique documents:        {len(docs)}")
        out(f"Rows with no target:     {len(unresolved)}")

        # 1. Unresolved rows (the "does not fetch all files" failures)
        if unresolved:
            out("\nMISS — index rows that resolved to zero PDFs:")
            for e in unresolved:
                out(f"  - [{e.get('year')}] {e['title'][:70]} "
                    f"({e.get('_landing_status', '?')}) {e['href']}")

        # 2. Download errors
        errors = [d for d in docs if d.get("error")]
        if downloaded:
            ok = [d for d in docs if d.get("filename") and not d.get("error")]
            out(f"\nDownloaded OK:           {len(ok)}")
            out(f"Download errors:         {len(errors)}")
            if errors:
                out("\nERROR — documents that failed to download cleanly:")
                for d in errors:
                    out(f"  - {d['error']}: {d.get('title', '')[:60]} "
                        f"[{d['download_url']}]")

        # 2a. File-type breakdown of what we actually have.
        def file_ext(d):
            fn = d.get("filename") or ""
            return (Path(fn).suffix.lower().lstrip(".") or "?") if fn else None
        by_ext = Counter(file_ext(d) for d in docs if file_ext(d))
        if by_ext:
            out("\nFile types:")
            for ext, n in by_ext.most_common():
                out(f"  {ext}: {n}")

        # 2b. Externally hosted documents (e.g. HAS) worth a manual double-check
        external = [d for d in docs
                    if urlparse(d["download_url"]).netloc not in
                    ("sfar.org", "www.sfar.org")]
        if external:
            out(f"\nExternally hosted documents (not on sfar.org): {len(external)}")
            for d in external:
                out(f"  - [{d.get('year')}] {d.get('title', '')[:55]} "
                    f"→ {urlparse(d['download_url']).netloc}")

        # 3. Year coverage
        by_year = Counter(str(d.get("year") or "unknown") for d in docs)
        out("\nPer-year document counts:")
        for y in sorted(by_year, reverse=True):
            out(f"  {y}: {by_year[y]}")
        expected_years = {str(y) for y in range(2000, 2027)}
        missing_years = sorted(expected_years - set(by_year))
        if missing_years:
            out(f"\n⚠️  Years with no documents: {', '.join(missing_years)}")
        if by_year.get("2026", 0) == 0:
            out("⚠️  2026 is empty — discovery likely incomplete.")

        # 4. Reference list cross-check
        ref_titles = self._reference_titles()
        if ref_titles:
            disc_norm = {self._norm_title(d["title"]) for d in docs}
            missing_ref = [t for t in ref_titles
                           if self._norm_title(t) not in disc_norm]
            out(f"\nReference titles:        {len(ref_titles)}")
            out(f"Reference titles missing from discovery: {len(missing_ref)}")
            for t in missing_ref[:40]:
                out(f"  - {t[:75]}")
            if len(missing_ref) > 40:
                out(f"  ... and {len(missing_ref) - 40} more")

        # 5. Superset vs curated 87
        if self.existing_pdfs and downloaded:
            got = {d["filename"].lower() for d in docs if d.get("filename")}
            still_missing = sorted(self.existing_pdfs - got)
            new_docs = sorted(got - self.existing_pdfs)
            out(f"\nCurated ../pdfs basenames:            {len(self.existing_pdfs)}")
            out(f"Curated basenames not re-downloaded:  {len(still_missing)}")
            out(f"Genuinely new basenames:              {len(new_docs)}")

        comprehensive = not unresolved and (not downloaded or not errors)
        out("\n" + "=" * 60)
        out("RESULT: " + ("✅ COMPREHENSIVE" if comprehensive
                          else "❌ NOT COMPREHENSIVE — see MISS/ERROR above"))
        out("=" * 60)

        report = self.output_dir / "completeness_report.txt"
        report.write_text("\n".join(lines), encoding="utf-8")
        print(f"\n🧾 Wrote {report}")
        return comprehensive

    @staticmethod
    def _norm_title(t):
        return re.sub(r"[^a-z0-9]+", " ", strip_accents(t.lower())).strip()

    def _reference_titles(self):
        """Parse recommendation titles from the reference markdown, best-effort."""
        if not REFERENCE_MD.exists():
            return []
        titles = []
        in_list = False
        for raw in REFERENCE_MD.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if line.startswith("## Chronologie"):
                in_list = True
                continue
            if not in_list:
                continue
            # stop at the trailing index/legend section
            if line.startswith("**Anesthésie") or "Index :" in line:
                break
            if not line or line.startswith(("#", "!", "|", "<", "*", "[", "-")):
                continue
            if re.match(r"^\d{4}$", line) or line.startswith("^"):
                continue
            # keep reasonably long prose lines (actual titles)
            if len(line) > 15:
                titles.append(line)
        return titles

    # -- orchestration ----------------------------------------------------- #

    def run(self, dry_run=False):
        entries = self.discover()
        if not entries:
            print("❌ No entries discovered.")
            return False

        docs, unresolved = self.resolve_all(entries)

        if not dry_run:
            self.download_all(docs)

        # Written after download so it carries the resulting filenames — the
        # next run reads these to skip files already on disk without re-fetching.
        self.write_discovery_json(docs, unresolved)
        self.write_manifest(docs)
        comprehensive = self.verify_complete(docs, unresolved, downloaded=not dry_run)
        return comprehensive


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main():
    parser = argparse.ArgumentParser(
        description="SFAR Recommendations PDF Scraper",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="Discover + resolve + manifest, but do not download.")
    parser.add_argument("--verify-only", action="store_true",
                        help="Re-run the completeness check on existing discovery.json.")
    parser.add_argument("--index-html", type=str, default=None,
                        help="Path to the index HTML snapshot "
                             "(default: 'Recommandations - La SFAR.html').")
    parser.add_argument("--live-index", action="store_true",
                        help="Fetch the recommendations index live instead of the snapshot.")
    parser.add_argument("--output", type=str, default=str(SCRIPT_DIR),
                        help="Output directory (default: script directory).")
    parser.add_argument("--workers", type=int, default=6,
                        help="Concurrent workers for resolve/download (default: 6).")
    parser.add_argument("--delay", type=float, default=0.5,
                        help="Minimum delay between live requests, seconds (default: 0.5).")
    args = parser.parse_args()

    scraper = SFARScraper(
        output_dir=Path(args.output).resolve(),
        index_html=args.index_html,
        live_index=args.live_index,
        workers=args.workers,
        delay=args.delay,
    )

    if args.verify_only:
        path = scraper.output_dir / "discovery.json"
        if not path.exists():
            print(f"ERROR: {path} not found; run discovery first.")
            sys.exit(1)
        data = json.loads(path.read_text(encoding="utf-8"))
        docs = data.get("documents", [])
        unresolved = data.get("unresolved", [])
        downloaded = any(d.get("filename") for d in docs)
        ok = scraper.verify_complete(docs, unresolved, downloaded=downloaded)
        sys.exit(0 if ok else 2)

    ok = scraper.run(dry_run=args.dry_run)
    sys.exit(0 if ok else 2)


if __name__ == "__main__":
    main()
